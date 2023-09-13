from openpyxl import load_workbook, Workbook
workbook = load_workbook(filename="./Data/Invoice UNS036 11-9-23 (1).xlsx", data_only=True)
alphabetical_file = "WCA-APC Masterlist Alphabetical [F].xlsx"
box_no_file = "WCA-APC Masterlist Box No. [F].xlsx"
curr_sheet = workbook.active

wca_items, apc_items, tc_items = [], [], []
store_items = []

def iterate_sheet(curr_sheet):
    curr_number = 0
    for row in curr_sheet.iter_rows(min_row=13, min_col=1, max_row=curr_sheet.max_row+1, max_col=9):
        is_main_wh = True
        curr_item = []
        for cell in row:
            

            if cell.column_letter == "E":
                #if our item description is None, that means we have reached the end of the invoice so break out of for loop
                if cell.value == None:
                    return

                plant_name = cell.value
                if 'Vesicularia montagnei' in plant_name :
                    plant_name = cell.value.replace("Vesicularia montagnei", "Christmas Moss")
                elif 'Vesicularia dubyana' in plant_name:
                    plant_name = cell.value.replace("Vesicularia dubyana", "Java Moss")
                elif 'Christmas Moss (Floating)' in plant_name:
                    plant_name = cell.value.replace("Christmas Moss (Floating)", "Floating Christmas Moss")
                elif "Portions" in plant_name:
                    plant_name = cell.value.replace("TC Portions", "TC")
                curr_item.append(plant_name)
                continue

            if cell.column_letter == "A":
                #if BOX No. in col A is None since only one instance of the BOX No. is listed in the invoice
                if cell.value == None:
                    curr_item.append(curr_number)
                    continue
                else:
                    curr_number = cell.value
                
            if cell.column_letter == "B":
                #if the ORDER NO. in col B is NONE, that means our current item is for main warehouse
                #if it is not NONE, that means our current item is a store order
                if cell.value == None:
                    continue
                else:  
                    is_main_wh = False

            curr_item.append(cell.value)

        if not is_main_wh:
            store_items.append(curr_item)
        if ' TC' in curr_item[3] or "Small Cup" in curr_item[3]:
            tc_items.append(curr_item)
        elif curr_item[1] == "U":
            wca_items.append(curr_item)
        elif curr_item[1] == "B":
            apc_items.append(curr_item)
    
def write_to_file(data, filename, wb):
    total_data = [["BOX No.", None, "CODE", "DESCRIPTION", "QUANTITY"]]
    for items in data:
        for item in items:
            total_data.append(item)
        total_data.append([' ']*len(items[0]))
    
    sheet = wb.active
    for row in range(len(total_data)):
        for col in range(len(total_data[0])):
            curr_cell = sheet.cell(row = row+1, column = col+1)
            curr_cell.value = total_data[row][col]
    wb.save(f"./Finalized Data/{filename}")


iterate_sheet(curr_sheet)
write_to_file([wca_items, apc_items, tc_items], box_no_file, Workbook())
#TODO sort by alphabetical

def print_items(items):
    for item in items:
        for value in item:
            print(value, end=" ")
        print()
    print()

print("==========================WCA==========================")
print_items(wca_items)
print("==========================APC==========================")
print_items(apc_items)
print("==========================TC==========================")
print_items(tc_items)
print("==========================STORE==========================")
#print_items(store_items)



    


